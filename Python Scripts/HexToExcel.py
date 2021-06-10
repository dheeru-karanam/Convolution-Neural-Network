import openpyxl
import UpdatedBtoD

row = 13
col = 13
f = open("2.hex","r") #Input Filename
FileName = "June7_Relu21_Kernel1_Relu2.xlsx" #Output Filename
work_book = openpyxl.Workbook()
work_book.save(FileName)
try:
    work_book = openpyxl.load_workbook(filename="June7_Relu21_Kernel1_Relu2.xlsx")
    sheet = work_book.active
except FileNotFoundError as e:
    print(e)
    exit()
red = openpyxl.styles.colors.Color(rgb='00FF0000')
fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=red)
for i in range(0,row):
    for j in range(0,col):
        char = f.readline().rstrip("\n")
        char = "0x"+ char
        num =  int(char, 16)
        char = UpdatedBtoD.BtoD(num,24,20)
        cell = sheet.cell(row=i+1,column=j+1)
        cell.value = char
        if(char != 0): #character you want to omit
            cell.fill = fill
        work_book.save(filename="June7_Relu21_Kernel1_Relu2.xlsx")

work_book.close()  
f.close()
